"""
report_common.py — Helper condivisi per i fogli Excel di dettaglio
(pivot per quarto, cross-table, situation hierarchy).

Usato sia da analyze.py (un foglio per play call) che da
analyze_players.py (un foglio per giocatore on-court).
"""

import re
from collections import Counter, defaultdict

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_COLOR = "1F3864"
ALT_COLOR    = "D9E1F2"


def pct(n, tot):
    return round(n / tot * 100, 1) if tot else 0


# ---------------------------------------------------------------------------
# Punti per possesso
# ---------------------------------------------------------------------------

_FT_RE         = re.compile(r"^(\d+)-(\d+)\s*FT$")       # "2-2 FT" -> 2 punti su 2 liberi
_SHOT_BONUS_RE = re.compile(r"^(\d)P\s*\+\s*(\d)$")      # "2P + 1" -> 2 (tiro) + 1 (libero bonus)
_SHOT_RE       = re.compile(r"^(IN|\dP)([+-])$")         # "IN+", "2P-", "3P+", ...


def result_points(result):
    """Ritorna (punti, è_possesso) per un valore RESULTS.

    in/2p/3p + = segnato (2 o 3 punti), - = sbagliato (0 punti)
    X-Y FT = X punti sui liberi (X segnati su Y)
    F e OB non sono possessi
    """
    r = (result or "").strip().upper()
    if not r or r in ("F", "OB"):
        return 0, False
    if r.startswith("TO"):
        return 0, True

    m = _FT_RE.match(r)
    if m:
        return int(m.group(1)), True

    m = _SHOT_BONUS_RE.match(r)
    if m:
        return int(m.group(1)) + int(m.group(2)), True

    m = _SHOT_RE.match(r)
    if m:
        base = 3 if m.group(1) == "3P" else 2
        return (base if m.group(2) == "+" else 0), True

    return 0, False


def ppp_stats(rows):
    """Punti totali, possessi e punti-per-possesso per una lista di righe."""
    points = possessions = 0
    for r in rows:
        p, is_poss = result_points(r.get("RESULTS"))
        if is_poss:
            points += p
            possessions += 1
    ppp = round(points / possessions, 2) if possessions else None
    return {"points": points, "possessions": possessions, "ppp": ppp}


def make_sheet(wb, name):
    return wb.create_sheet(name)


def autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def get_situations(row):
    """Espande valori multipli separati da virgola: 'OR, P&R Handler' → ['OR', 'P&R Handler']"""
    val = row.get("SITUATION", "")
    return [s.strip() for s in val.split(",")] if val else []


SITUATION_FAMILIES = {
    "P&R":        lambda s: s.startswith("P&R"),
    "1on1":       lambda s: "1on1" in s,
    "OR":         lambda s: "OR" in s,
    "FB":         lambda s: s == "FB",
    "POST UP":    lambda s: "POST" in s,
    "HO":         lambda s: s.startswith("HO"),
    "SCREENS":    lambda s: s.startswith("SCREENS"),
    "Altro":      lambda s: True,  # catch-all
}


def situation_hierarchy(rows):
    """Ritorna righe per la sezione gerarchia situation."""
    N = len(rows)
    if not N:
        return []

    out = [("--- SITUATION BREAKDOWN ---", "N", "%", "Results (top3)", "Qual. media", "Paint Touch %")]
    out.append(("", "", "", "", "", ""))

    def family_of(s):
        for family, match_fn in SITUATION_FAMILIES.items():
            if match_fn(s):
                return family

    for family in SITUATION_FAMILIES:
        # tutte le righe con almeno una situation di questa famiglia
        # (una riga con più situation di famiglie diverse conta in ciascuna)
        fam_rows = [r for r in rows if any(family_of(s) == family for s in get_situations(r))]
        if not fam_rows:
            continue

        fam_pct = pct(len(fam_rows), N)
        out.append((family, len(fam_rows), fam_pct, "", "", ""))

        # specifiche situations dentro la famiglia
        spec_counter = Counter(
            s for r in fam_rows for s in get_situations(r) if family_of(s) == family
        )
        for sit, cnt in sorted(spec_counter.items(), key=lambda x: -x[1]):
            sit_rows = [r for r in fam_rows if sit in get_situations(r)]
            NS = len(sit_rows)

            # top 3 results
            res_c = Counter(r["RESULTS"] for r in sit_rows if r.get("RESULTS"))
            top_res = ", ".join(f"{k}:{v}" for k, v in res_c.most_common(3))

            # quality media
            qvals = [int(r["QUALITY"]) for r in sit_rows if r.get("QUALITY")]
            avg_q = round(sum(qvals) / len(qvals), 1) if qvals else ""

            # paint touch %
            pt_n = sum(1 for r in sit_rows if r.get("PAINT TOUCHES"))
            pt_pct = pct(pt_n, NS)

            out.append((f"  {sit}", NS, pct(NS, N), top_res, avg_q, pt_pct))

        out.append(("", "", "", "", "", ""))

    return out


def cross_table(rows, col_a, col_b, label_a, label_b, get_a=None, get_b=None):
    """
    Incrocio tra due dimensioni.
    get_a/get_b: funzioni opzionali per estrarre lista di valori da una riga
    (es. per SITUATION con virgole). Default: valore singolo da col.
    """
    if get_a is None:
        get_a = lambda r: [r[col_a]] if r.get(col_a) else []
    if get_b is None:
        get_b = lambda r: [r[col_b]] if r.get(col_b) else []

    # raggruppa: per ogni valore di A, conta i valori di B
    groups = defaultdict(list)
    for row in rows:
        for a in get_a(row):
            for b in get_b(row):
                groups[a].append(b)

    if not groups:
        return []

    out = [("", "", ""), (f"--- {label_a} × {label_b} ---", "N", "%")]
    for a_val, b_vals in sorted(groups.items(), key=lambda x: -len(x[1])):
        n_a = len(b_vals)
        out.append((a_val, n_a, ""))
        for b_val, cnt in sorted(Counter(b_vals).items(), key=lambda x: -x[1]):
            out.append((f"  {b_val}", cnt, pct(cnt, n_a)))
    return out


PERIODS = [("Totale", None), ("Q1", "1 Q"), ("Q2", "2 Q"), ("Q3", "3 Q"), ("Q4", "4 Q"), ("CT", "CT")]


def period_rows(all_rows, quarter_val):
    if quarter_val is None:
        return all_rows
    return [r for r in all_rows if r.get("QUARTER") == quarter_val]


def pivot_section(label, all_rows, get_vals):
    """Genera righe pivot: colonne = Totale, Q1, Q2, Q3, Q4, CT."""
    # Raccoglie tutti i valori unici
    all_vals = sorted(set(v for r in all_rows for v in get_vals(r)))
    if not all_vals:
        return []

    # Header sezione
    header = [f"--- {label} ---"] + [c for p, _ in PERIODS for c in (p + " N", p + " %")]
    rows_out = [tuple(header)]

    for val in all_vals:
        row = [val]
        for p_label, q_val in PERIODS:
            subset = period_rows(all_rows, q_val)
            n_period = len(subset)
            n_val = sum(1 for r in subset if val in get_vals(r))
            row += [n_val, pct(n_val, n_period)]
        rows_out.append(tuple(row))

    rows_out.append(("",) * len(header))
    return rows_out


def detail_sheet(wb, name, rows):
    """Crea un foglio con pivot per quarto + cross-table + situation hierarchy
    per l'insieme di righe passato (es. azioni di un play call, oppure tutte
    le azioni in cui un giocatore è in campo)."""
    safe_name = name[:31]
    ws = make_sheet(wb, safe_name)

    col_headers = [""] + [c for p, _ in PERIODS for c in (f"{p} N", f"{p} %")]
    for col, h in enumerate(col_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center")

    all_data = []

    # Totale azioni per periodo (riga intro)
    intro = ["Azioni"]
    for p_label, q_val in PERIODS:
        n = len(period_rows(rows, q_val))
        intro += [n, ""]
    all_data.append(tuple(intro))

    # PPP (punti per possesso) per periodo
    ppp_row  = ["PPP"]
    pts_row  = ["Punti"]
    poss_row = ["Possessi"]
    for p_label, q_val in PERIODS:
        stats = ppp_stats(period_rows(rows, q_val))
        ppp_row  += [stats["ppp"] if stats["ppp"] is not None else "", ""]
        pts_row  += [stats["points"], ""]
        poss_row += [stats["possessions"], ""]
    all_data.append(tuple(ppp_row))
    all_data.append(tuple(pts_row))
    all_data.append(tuple(poss_row))

    all_data.append(("",) * len(col_headers))

    # Sezioni metriche
    all_data += pivot_section("RESULTS",       rows, lambda r: [r["RESULTS"]] if r.get("RESULTS") else [])
    all_data += pivot_section("SITUATION",     rows, get_situations)
    all_data += pivot_section("SHOT LOCATION", rows, lambda r: [r["Shot Location"]] if r.get("Shot Location") else [])
    all_data += pivot_section("QUALITY SHOT",  rows, lambda r: [r["QUALITY"]] if r.get("QUALITY") else [])
    all_data += pivot_section("O COVERAGES",   rows, lambda r: [r["O COVERAGES"]] if r.get("O COVERAGES") else [])
    all_data += pivot_section("PRESSING",      rows, lambda r: [r["PRESSING"]] if r.get("PRESSING") else [])

    # Paint touch: Senza + categorie
    def get_pt_ext(r):
        return [r["PAINT TOUCHES"]] if r.get("PAINT TOUCHES") else ["Senza"]
    all_data += pivot_section("PAINT TOUCH", rows, get_pt_ext)

    # Broken play
    def get_bp(r):
        return ["Broken Play"] if r.get("PLAN/BROKEN PLAY") else ["Non Broken"]
    all_data += pivot_section("BROKEN PLAY", rows, get_bp)

    # Legami (3 colonne, sezione separata)
    legami_data = []
    get_sit = get_situations
    get_res = lambda r: [r["RESULTS"]] if r.get("RESULTS") else []
    get_pt  = lambda r: [r["PAINT TOUCHES"]] if r.get("PAINT TOUCHES") else []
    get_qs  = lambda r: [r["QUALITY"]] if r.get("QUALITY") else []
    get_cov = lambda r: [r["O COVERAGES"]] if r.get("O COVERAGES") else []

    legami_data += cross_table(rows, "SITUATION", "RESULTS",       "SITUATION",     "RESULTS",      get_sit, get_res)
    legami_data += cross_table(rows, "SITUATION", "PAINT TOUCHES", "SITUATION",     "PAINT TOUCHES",get_sit, get_pt)
    legami_data += cross_table(rows, "SITUATION", "QUALITY",       "SITUATION",     "QUALITY SHOT", get_sit, get_qs)
    legami_data += cross_table(rows, "O COVERAGES","RESULTS",      "O COVERAGES",   "RESULTS",      get_cov, get_res)
    legami_data += cross_table(rows, "PAINT TOUCHES","RESULTS",    "PAINT TOUCHES", "RESULTS",      get_pt,  get_res)

    # Scrivi pivot (multi-colonna)
    for i, row in enumerate(all_data):
        is_alt = i % 2 == 1
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=2 + i, column=col, value=val)
            if is_alt:
                cell.fill = PatternFill("solid", fgColor=ALT_COLOR)

    # Scrivi legami affianco (colonna separata dopo gap)
    gap_col = len(col_headers) + 2
    for i, row in enumerate(legami_data):
        fill = PatternFill("solid", fgColor=ALT_COLOR) if i % 2 == 1 else None
        for col, val in enumerate(row, gap_col):
            cell = ws.cell(row=2 + i, column=col, value=val)
            if fill:
                cell.fill = fill
        if row and row[0].startswith("---"):
            for col in range(gap_col, gap_col + 3):
                ws.cell(row=2 + i, column=col).font = Font(bold=True, color="FFFFFF")
                ws.cell(row=2 + i, column=col).fill = PatternFill("solid", fgColor=HEADER_COLOR)

    # Situation hierarchy — colonna dopo legami
    # colonne: label(0), N(1), %(2), Results(3), Qual(4), PT%(5)
    hier_col = gap_col + 5
    hier_data = situation_hierarchy(rows)
    for i, row in enumerate(hier_data):
        is_alt = i % 2 == 1
        for offset, val in enumerate(row):
            col = hier_col + offset
            cell = ws.cell(row=2 + i, column=col, value=val)
            if is_alt:
                cell.fill = PatternFill("solid", fgColor=ALT_COLOR)
        if row and str(row[0]).startswith("---"):
            for col in range(hier_col, hier_col + 6):
                ws.cell(row=2 + i, column=col).font = Font(bold=True, color="FFFFFF")
                ws.cell(row=2 + i, column=col).fill = PatternFill("solid", fgColor=HEADER_COLOR)
        elif row and row[0] and not str(row[0]).startswith("  "):
            for col in range(hier_col, hier_col + 6):
                ws.cell(row=2 + i, column=col).font = Font(bold=True)

    autofit(ws)
